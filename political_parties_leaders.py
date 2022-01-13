import datetime
import json
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup as bS
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font
import database_operations as dc
import email_utility as eu


db = dc.connect_db()
curr_timestamp = datetime.datetime.now().replace(microsecond=0)
today = str(datetime.date.today())  # - datetime.timedelta(days=1))
url = "https://www.cia.gov/the-world-factbook/page-data/field/political-parties-and-leaders/page-data.json"

wb = Workbook()
wb1 = Workbook()
email_for_manual_entries = 'us@yopmail.com'
emails_for_pep_data = ['abc_ocr@yopmail.com', 'bcd_orc@yopmail.com', 'cde_ocr@yopmail.com']
manual_entries_filename = 'manual_entries.xlsx'
pep_dest_filename = 'pep/political_parties_and_leaders.xlsx'
ws1 = wb.active
ws1.title = "Full List of all PEP"
ws2 = wb.create_sheet("Diff (Previous Day Changes)")
ws3 = wb1.active
ws3.title = 'Manual Entries'
manual = "MANUAL"
automated = "AUTOMATED"


def process_sheet1():
    result_set = dc.execute_select_query("SELECT country, party_name, leader_name FROM pep where deleted_on = null", ())
    country_list = list()
    party_list = list()
    leader_list = list()
    for result in result_set:
        country_list.append(result[0])
        party_list.append(result[1])
        leader_list.append(result[2])
    data = {'Country': country_list, 'Party_Name': party_list, 'Leader_Name': leader_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws1)


def fetch_row(import_id):
    return dc.get_entity("SELECT row from pep_import_history where id = %s", (import_id,))


def process_sheet2():
    result_set = dc.execute_select_query("SELECT country, party_name, leader_name, import_id, deleted_on FROM pep", ())
    country_list = list()
    party_list = list()
    leader_list = list()
    row_list = list()
    action_list = list()
    for result in result_set:
        country_list.append(result[0])
        party_list.append(result[1])
        leader_list.append(result[2])
        import_id = result[3]
        text_as_is = fetch_row(import_id)
        row_list.append(text_as_is[0])
        if result[4] is not None:
            action_list.append("DELETED")
        else:
            action_list.append("ADDED")
    data = {'Text_As_Is': row_list, 'Country': country_list, 'Party_Name': party_list,
            'Leader_Name': leader_list, 'Action': action_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws2)
    subject = 'PEP data'
    content = 'Hello Team,\n\nAttached excel contains PEP data in 2 versions' \
              '\n\n\n   1. Full List of all PEP\n   2. Diff (Previous Day Changes)'
    wb.save(filename=pep_dest_filename)
    eu.initiate_email(emails_for_pep_data, subject, content, pep_dest_filename)


def generate_excel_from_dataframe(df, work_sheet, highlight=False):
    work_sheet.append(tuple(df.keys()))
    header_len = len(tuple(df.keys()))
    for i in range(1, header_len + 1):
        cell = chr(i + 64) + str(1)
        work_sheet[cell].font = Font(bold=True)
    if not highlight:
        for tup in df.itertuples(index=False):
            work_sheet.append(tup)
    else:
        my_red = Color(rgb='00FFCCCB')
        my_fill = PatternFill(patternType='solid', fgColor=my_red)
        for index, tup in enumerate(df.itertuples(index=False)):
            row = tup[1]
            work_sheet.append(tup)
            if not ("[" in row and "]" in row):
                for i in range(1, header_len+1):
                    cell = chr(i + 64) + str(index+2)
                    work_sheet[cell].fill = my_fill


def notify_manual_entries_to_sme():
    result_set = dc.execute_select_query("SELECT country, row, is_note FROM pep_import_history WHERE type = 'MANUAL'", ())
    country_list = list()
    row_list = list()
    is_note_list = list()
    for result in result_set:
        country_list.append(result[0])
        row_list.append(result[1])
        is_note_list.append('Yes' if result[2] == 1 else 'No')
    data = {'Country': country_list, 'Row': row_list, 'Is_It_A_Note': is_note_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws3, True)
    subject = "Manual Entries in PEP List"
    content = "Hello SME Team, \n\nAttached excel contains the manual entries to be filtered, " \
              "validated & inserted into the pep table in database."
    wb1.save(filename=manual_entries_filename)
    eu.initiate_email(email_for_manual_entries, subject, content, manual_entries_filename)


def clean_text_with_bs(text):
    if not text.endswith(">"):
        remaining_text = text.split(">")[len(text.split(">")) - 1]
        html_data = bS(text, "lxml-html")
        if html_data.text == remaining_text or remaining_text in html_data.text:
            return_text = html_data.text.strip()
        else:
            return_text = (html_data.text + remaining_text).strip()
    else:
        html_data = bS(text, "lxml-xml")
        return_text = html_data.text.strip()
    if return_text == "":
        return text.replace("<strong>", "").strip()
    else:
        return return_text


# call the URL & collect response
def fetch_data_from_url():
    file = urllib.request.urlopen(url)
    data = file.read()
    file.close()
    return str(data, "UTF-8")


# store the response into an XML file in string format
def do_file_operations():
    # file_name = "json_files/2021-11-19.json"
    file_name = "json_files/" + today + ".json"
    f = open(file_name, "w")
    f.write(fetch_data_from_url())
    f.close()
    return file_name


def check_for_keywords(data_string):
    select_keywords_qry = "SELECT * from pep_import_keywords where keyword = %s and is_ignored = %s"
    return dc.check_for_duplication(select_keywords_qry, (data_string, 1))


def get_designations():
    designations = dc.execute_select_query("SELECT designation from pep_designations where 1", ())
    return designations


def check_for_designation_string(leader):
    designations = get_designations()
    des_list = list()
    for designation in designations:
        des_list.append(designation[0])
    for designation in des_list:
        leader = leader.replace(designation, "").strip()
    return leader


def store_pep_history(country_name, row, entry_type, is_note, created_on, updated_on, deleted_on):
    pep_hist_insert = "INSERT INTO pep_import_history" \
                      "(country, row, type, is_note, created_on, modified_on, deleted_on) " \
                      "VALUES (%s, %s, %s, %s, %s, %s, %s)"
    duplication_qry = "SELECT * from pep_import_history where country = %s and row = %s and type = %s and is_note = %s"
    insert_values = (country_name, row, entry_type, is_note, created_on, updated_on, deleted_on)
    check_values = (country_name, row, entry_type, is_note)
    if not dc.check_for_duplication(duplication_qry, check_values):
        dc.execute_query_single(pep_hist_insert, insert_values)


# def check_comma_after_bracket(data_string):
#     if "]" in data_string:
#         splits = data_string.split("]")
#         for split in splits:
#             if split.startswith(","):
#                 return True
#     else:
#         return False


def process_pep_history(country_name, data_string, mark_as_deleted):
    today_date = datetime.date.today()
    created_on = today_date - datetime.timedelta(days=1) if mark_as_deleted else today_date
    deleted_on = today_date if mark_as_deleted else None
    if "<" in data_string:
        data_string = clean_text_with_bs(data_string.replace("<p>", "").replace("</p>", "")).replace("</strong>", "").strip()
    left_bracket_count = data_string.count('[')
    right_bracket_count = data_string.count(']')
    if left_bracket_count == 1 and right_bracket_count == 1:
        if data_string.split("]")[1] != "" and data_string.split("]")[1].startswith(")"):
            store_pep_history(country_name, data_string, manual, False, created_on, None, deleted_on)
        elif data_string.startswith("note:") or data_string.startswith("note :"):
            is_note = True if check_for_keywords(data_string) else False
            store_pep_history(country_name, data_string, manual, is_note, created_on, None, deleted_on)
        elif data_string.startswith("one registered party: ") and data_string.replace("one registered party: ", "") != "":
            store_pep_history(country_name, data_string.replace("one registered party: ", ""), automated, False, created_on, None, deleted_on)
        elif data_string.startswith("other: ") and data_string.replace("other: ", "") != "":
            store_pep_history(country_name, data_string.replace("other: ", ""), automated, False, created_on, None, deleted_on)
        elif data_string.startswith("New - ") and data_string.replace("New - ", "") != "":
            store_pep_history(country_name, data_string.replace("New - ", ""), automated, False, created_on, None, deleted_on)
        elif data_string.split("[")[1].replace("]", "").startswith("collective leadership") \
                or data_string.split("[")[1].replace("]", "") == "NA" \
                or data_string.split("[")[1].replace("]", "") == "N/A"\
                or data_string.split("[")[1].replace("]", "") == "joint leadership of several medical doctors":
            pass
        else:
            store_pep_history(country_name, data_string, automated, False, created_on, None, deleted_on)
    else:
        if data_string.strip() != "" and data_string.strip() != "(" and (not (data_string.startswith("<") and data_string.endswith(">"))):
            is_note = True if check_for_keywords(data_string) else False
            store_pep_history(country_name, data_string, manual, is_note, created_on, None, deleted_on)
        # elif check_comma_after_bracket(data_string): # case for Uruguay 1st merged entry
        #     for split in data_string.split("],"):
        #         store_pep_history(country_name, split, automated, False, today_date, today_date, None)


def fetch_pep_history():
    country_dict = dict()
    result_set = dc.execute_select_query("SELECT * from pep_import_history", ())
    country_name = ''
    for result in result_set:
        import_id = result[0]
        country = result[1]
        if country not in country_dict.keys():
            country_dict[country] = list()
        row = result[2]
        is_note = result[3]
        entry_type = result[4]
        created_on = result[5]
        deleted_on = result[7]
        data_tup = (import_id, row, is_note, entry_type, created_on, deleted_on)
        if country_name != country:
            country_dict[country].append(data_tup)
            country_name = country
        else:
            country_dict[country].append(data_tup)
    # old_df = pd.DataFrame(old_dict)
    # print(old_df.keys())

    return country_dict

    # for country in old_df.keys():
    #     print(country)
    #     print(old_df.get(country))
    #     print("============================================")
    # for tup in old_df.iteritems():
    #     print(tup)
    # print(old_df)


def process_data(data_string):
    note = ""
    if "[" in data_string:
        party_name = data_string.split("[")[0]
        party_leader = data_string.split("[")[1]
        if "]" not in party_leader:
            party_leader += "]"
        if not party_leader.split("]")[1] == "":
            note = party_leader.split("]")[1].strip()
            party_leader = party_leader.replace(note, "").strip()
            if note.startswith("<br><br>"):
                note = ""
    elif "(" in data_string:
        party_name = data_string.split("(")[0]
        party_leader = ""
        note = "(" + data_string.split("(")[1]
        # splitted_data = note.replace("(", "").replace(")", "").split(" ")
        if note.replace("(", "").replace(")", "").isupper():  # for Azerbaijan -> (PDR) scenario
            party_name += note
            note = ""
        # elif len(splitted_data) > 1 and splitted_data[
        #     1].isupper():  # for scenarios where party leaders names are in () parenthesis & not in []
        #     party_leader = note.replace("(", "").replace(")", "")
        #     note = ""
    elif "; note" in data_string or ";note" in data_string:
        if "; note" in data_string:
            party_name = data_string.split("; note")[0]
            party_leader = ""
            note = data_string.split("; note")[1]
        else:
            party_name = data_string.split(";note")[0]
            party_leader = ""
            note = data_string.split(";note")[1]
    else:
        party_name = data_string
        party_leader = ""
        note = ""
    data_tuple = (party_name.strip(),
                  party_leader.replace("]", "").replace(note.strip(), "").strip(),
                  note.lstrip(";").strip())
    return data_tuple


def clean_note_string(note):
    if note == "," or note == ")" or note == "(":
        note = ""
    return note.replace("&nbsp;", " ").replace("<p>", "").replace("</p>", "").replace("<br>", "") \
        .replace("; note", "").replace(";note", "").replace("&amp;", "&").replace("&rsquo;", "’") \
        .replace("<strong>", "").replace("</strong>", "").strip().lstrip("-").lstrip(";")


def clean_party_name_string(party_name):
    return party_name.replace("&amp;", "&").replace("&nbsp;", "").replace("&rsquo;", "’").strip()


def clean_party_leader_string(party_leader):
    if party_leader.endswith(")"):
        party_leader = party_leader.replace(")", "").strip()
    return party_leader.replace("&amp;", "&").replace("&nbsp;", " ").replace("&rsquo;", "’")\
        .replace("&aacute;", "á").replace(" (both claiming leadership)", "").strip()


def parse_json_file():
    file_name = "json_files/2021-12-03.json"
    f = open(file_name,)
    json_data = json.load(f)
    political_parties_json_string = json_data["result"]["data"]["page"]["json"].replace("\"", '"')
    pp_json_data = json.loads(political_parties_json_string)
    for country_data in pp_json_data["countries"]:
        country_name = country_data["name"].strip()
        parties_string = country_data["data"].replace("&amp;", "&").replace("&nbsp;", " ").replace("&rsquo;", "’").strip()
        if "<br />" in parties_string:
            parties = parties_string.split("<br />")
            for party in parties:
                if party != "":
                    party = party.replace("[[", "[").replace("]]", "]").strip()
                    if "<br>" in party:
                        if "<br><br>" in party:
                            first_part = party.split("<br><br>")[0]
                            second_part = party.split("<br><br>")[1]
                        else:
                            first_part = party.split("<br>")[0]
                            second_part = party.split("<br>")[1]
                        if first_part.strip() != "":
                            process_pep_history(country_name, first_part, False)
                        if second_part.strip() != "":
                            process_pep_history(country_name, second_part, False)
                    else:
                        process_pep_history(country_name, party, False)
        elif "<br>" in parties_string:
            parties = parties_string.split("<br />")
            party = parties[0].replace("[[", "[").replace("]]", "]").strip()
            if party != "":
                if "<br>" in party:
                    if "<br><br>" in party:
                        first_part = party.split("<br><br>")[0]
                        second_part = party.split("<br><br>")[1]
                    else:
                        first_part = party.split("<br>")[0]
                        second_part = party.split("<br>")[1]
                    if first_part.strip() != "":
                        process_pep_history(country_name, first_part, False)
                    if second_part.strip() != "":
                        process_pep_history(country_name, second_part, False)
                else:
                    process_pep_history(country_name, party, False)
        else:
            parties_string = parties_string.replace("[[", "[").replace("]]", "]").strip()
            if parties_string != "":
                process_pep_history(country_name, parties_string, False)


def process_automated_peps():
    result_set = dc.execute_select_query("SELECT * FROM pep_import_history WHERE type = 'AUTOMATED'", ())
    for result in result_set:
        import_id = result[0]
        country = result[1]
        row = result[2]
        deleted_on = result[7]
        data_tuple = process_data(row)
        party = data_tuple[0]
        party_leaders = data_tuple[1].replace("coalition led by ", "").replace("associated with former ", "")\
            .replace("Central Committee", "").replace("former PM ", "").replace("Honorary Chairman", "").strip()
        note = data_tuple[2]
        if "+" in party_leaders:
            party_leader_splitted = party_leaders.split("+")
            for leader in party_leader_splitted:
                leader_name = check_for_designation_string(leader.strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, None, note, automated, import_id, deleted_on)
        elif "," in party_leaders:
            party_leader_splitted = party_leaders.split(",")
            for index, leader in enumerate(party_leader_splitted):
                if " and " in leader:
                    party_leader_splitted1 = leader.split(" and ")
                    for leader1 in party_leader_splitted1:
                        leader_name = check_for_designation_string(leader1.strip()).lstrip("-")
                        if leader_name != "" and leader_name.strip() != "Dr." and leader_name != "Jr.":
                            do_entry_in_pep(country, party, leader_name, None, note, automated, import_id, deleted_on)
                else:
                    if not leader.strip().startswith("aka"):
                        leader_alias = None
                        if index <= 0:
                            if party_leader_splitted[index+1] is not None:
                                if party_leader_splitted[index+1].strip().startswith("aka"):
                                    leader_alias = party_leader_splitted[index+1].strip().replace("aka", "")
                            leader_name = check_for_designation_string(leader.strip()).lstrip("-")
                            if leader_name != "" and leader_name.strip() != "Dr." and leader_name.strip() != "Jr.":
                                do_entry_in_pep(country, party, leader_name, leader_alias, note, automated, import_id, deleted_on)
                        else:
                            leader_name = check_for_designation_string(leader.strip()).lstrip("-")
                            if leader_name != "" and leader_name.strip() != "Dr." and leader_name.strip() != "Jr.":
                                do_entry_in_pep(country, party, leader_name, leader_alias, note, automated, import_id, deleted_on)
        elif " and " in party_leaders:
            party_leader_splitted = party_leaders.split(" and ")
            for leader in party_leader_splitted:
                leader_name = check_for_designation_string(leader.replace("MPs", "").strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, None, note, automated, import_id, deleted_on)
        elif ";" in party_leaders:
            party_leader_splitted = party_leaders.split(";")
            for leader in party_leader_splitted:
                leader_name = check_for_designation_string(leader.strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, None, note, automated, import_id, deleted_on)
        elif " or " in party_leaders:
            party_leader_splitted = party_leaders.split(" or ")
            if len(party_leader_splitted) == 2:
                leader = party_leader_splitted[0]
                leader_alias = party_leader_splitted[1]
                leader_name = check_for_designation_string(leader.strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, leader_alias, note, automated, import_id, deleted_on)
        else:
            leader_name = check_for_designation_string(party_leaders.strip())
            if leader_name != "":
                do_entry_in_pep(country, party, leader_name, None, note, automated, import_id, deleted_on)


def do_entry_in_pep(country_name, party_name, leaders, leader_alias, notes, entry_type, import_id, deleted_on):
    check_duplication_qry = "SELECT * from pep where country = %s and party_name = %s and leader_name = %s " \
                            "and notes = %s and type = %s and import_id = %s"
    insertion_qry = "INSERT INTO pep(country, party_name, leader_name, leader_alias, notes, type, import_id, " \
                    "created_on, modified_on, deleted_on) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    party_name = clean_party_name_string(party_name)
    leaders = clean_party_leader_string(leaders)
    notes = clean_note_string(notes)
    today_date = datetime.date.today()
    created_on = today_date - datetime.timedelta(days=1) if deleted_on is not None else today_date
    if not (party_name == "" and leaders == "" and notes == ""):
        dup_val = (country_name, party_name, leaders, notes, entry_type, import_id)
        if not dc.check_for_duplication(check_duplication_qry, dup_val):
            insert_val = (country_name, party_name, leaders, leader_alias, notes, entry_type, import_id, created_on, None, deleted_on)
            dc.execute_query_single(insertion_qry, insert_val)


def clear_prev_data():
    dc.execute_query_single("delete from pep_import_history", ())
    dc.execute_query_single("delete from pep", ())


def check_row_deletion(country_data, new_data):
    for data in new_data:  # (A, B, C)
        if data[0] == country_data[0]:  # D
            return False
    return True


def check_row_addition(country_data, prev_data):
    for data in prev_data:
        if data[0] == country_data[0]:
            return True
    return False


def update_pep_import_entry(import_id, created_on, deleted_on):
    update_sql = "UPDATE pep_import_history SET created_on=%s, deleted_on=%s WHERE id=%s"
    update_val = (created_on, deleted_on, import_id)
    dc.execute_query_single(update_sql, update_val)


def process_pep_changes(prev_data, new_data):
    prev_countries = prev_data.keys()
    new_countries = new_data.keys()
    if len(prev_countries) == len(new_countries):
        for country in prev_countries:
            print(country)
            if len(prev_data[country]) > len(new_data[country]):  # some rows got deleted from recent data
                for country_data in prev_data[country]:
                    if check_row_deletion(country_data, new_data[country]):
                        process_pep_history(country, country_data[1], True)
            elif len(prev_data[country]) < len(new_data[country]):  # some rows got added to recent data
                for country_data in new_data[country]:
                    if check_row_addition(country_data, prev_data[country]):
                        process_pep_history(country, country_data[1], False)
            else:
                for country_data in new_data[country]:
                    country_data1 = prev_data[country]
                    import_id = country_data[0]
                    new_created_date = country_data1[3]
                    new_deleted_date = country_data1[4]
                    update_pep_import_entry(import_id, new_created_date, new_deleted_date)




# do_file_operations(url)
# parse_json_file()
prev_data = fetch_pep_history()  # fetch the old pep_history as {'country_name': [(data_tuple1), (data_tuple2), ...]}
clear_prev_data()  # as we have old data stored in our program clear it from database
parse_json_file()  # parse the recently downloaded JSON file to get new pep_import_history entries insert in DB
new_data = fetch_pep_history()  # fetch the new pep_history as {'country_name': [(data_tuple1), (data_tuple2), ...]}
process_pep_changes(prev_data, new_data)  # find diff b/w old & new data & mark lost entries as deleted in pep_history
# process_automated_peps()  #  now the pep data will be parsed according to new pep_history + deleted marked if any
notify_manual_entries_to_sme()
# process_sheet1()
# process_sheet2()  # while generating excel it'll see if there's any deleted entry in pep, mark it as DELETED in action
