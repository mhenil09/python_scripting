import datetime
import json
import urllib.request
from bs4 import BeautifulSoup as bS
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Color, PatternFill, Font
import database_operations as dc
import time


curr_timestamp = datetime.datetime.now().replace(microsecond=0)
today = datetime.date.today() # - datetime.timedelta(days=1)
url = "https://www.cia.gov/page-data/resources/world-leaders/foreign-governments/page-data.json"
base_url = "https://www.cia.gov/page-data"
suffix_url = "page-data.json"
file_name_prefix = "json_files/world_leaders/"
file_name_suffix = "_" + str(today) + ".json"

wb = Workbook()
dest_filename = 'world_leaders.xlsx'
ws = wb.active
ws.title = "World Leaders"
country_dict = dict()


def get_current_timestamp():
    ts = time.time()
    return datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')


# call the URL & collect response
def fetch_data_from_url(url):
    file = urllib.request.urlopen(url)
    data = file.read()
    file.close()
    return str(data, "UTF-8")


# store the response into an XML file in string format
def do_file_operations(url):
    file_name = "json_files/world_leaders/world_leaders.json"
    f = open(file_name, "w")
    f.write(fetch_data_from_url(url))
    f.close()
    return file_name

# do_file_operations(url)


def clean_ignore_words(text):
    replace_list = ["(Acting)", "(Interim)", "(Ret.)"]
    for replace in replace_list:
        text = text.replace(replace, "")
    if "," in text:
        comma_separated = text.split(",")
        before_comma = comma_separated[0]
        return before_comma
    return text


def clean_text(text):
    return text.replace("&#038;", "&").replace("&amp;", "&").replace("&#8217;", "’").strip()


def generate_excel_from_dataframe(df, work_sheet, highlight=False):
    work_sheet.append(tuple(df.keys()))
    header_len = len(tuple(df.keys()))
    for i in range(1, header_len + 1):
        cell = chr(i + 64) + str(1)
        work_sheet[cell].font = Font(bold=True)
    if not highlight:
        for tup in df.itertuples(index=False):
            work_sheet.append(tup)
    else:
        my_red = Color(rgb='00FFCCCB')
        my_fill = PatternFill(patternType='solid', fgColor=my_red)
        for index, tup in enumerate(df.itertuples(index=False)):
            row = tup[1]
            work_sheet.append(tup)
            if not ("[" in row and "]" in row):
                for i in range(1, header_len+1):
                    cell = chr(i + 64) + str(index+2)
                    work_sheet[cell].fill = my_fill


def parse_json_file():
    file_name = "json_files/world_leaders/world_leaders.json"
    f = open(file_name,)
    json_data = json.load(f)
    data = []
    world_leaders_json_string = json_data["result"]["data"]["governments"]["edges"]
    for node in world_leaders_json_string:
        node_data = json.loads(str(node).replace("'", '"'))
        country_name = node_data["node"]["title"]
        json_url = base_url + node_data["node"]["path"] + suffix_url
        country_dict[country_name] = json_url
        data_list = process_country_data(clean_text(country_name), json_url)
        for data1 in data_list:
            data.append(data1)
    return convert_to_pep_import_df(data)


def parse_json_file_to_insert():
    file_name = "json_files/world_leaders/world_leaders.json"
    f = open(file_name,)
    json_data = json.load(f)
    # data = []
    world_leaders_json_string = json_data["result"]["data"]["governments"]["edges"]
    for node in world_leaders_json_string:
        node_data = json.loads(str(node).replace("'", '"'))
        country_name = node_data["node"]["title"]
        json_url = base_url + node_data["node"]["path"] + suffix_url
        country_dict[country_name] = json_url
        process_country_data_to_insert(clean_text(country_name), json_url)


def process_country_data(country_name, json_url):
    file_name = file_name_prefix + country_name + file_name_suffix
    # f = open(file_name, "w")
    # f.write(fetch_data_from_url(json_url))
    # f.close()
    data = []
    f = open(file_name,)
    if f is not None:
        json_data = json.load(f)
        govt_count = len(json_data["result"]["data"]["page"]["acf"]["blocks"])
        if govt_count == 1:
            content = json_data["result"]["data"]["page"]["acf"]["blocks"][0]["free_form_content"]["content"]
            data_list = process_content(content, country_name)
            for data1 in data_list:
                data.append(data1)
        else:
            for i in range(0, govt_count):
                content = json_data["result"]["data"]["page"]["acf"]["blocks"][i]["free_form_content"]["content"]
                data_list = process_content(content, country_name)
                for data1 in data_list:
                    data.append(data1)
    return data


def process_country_data_to_insert(country_name, json_url):
    file_name = file_name_prefix + country_name + file_name_suffix
    # f = open(file_name, "w")
    # f.write(fetch_data_from_url(json_url))
    # f.close()
    # data = []
    f = open(file_name,)
    if f is not None:
        json_data = json.load(f)
        govt_count = len(json_data["result"]["data"]["page"]["acf"]["blocks"])
        if govt_count == 1:
            content = json_data["result"]["data"]["page"]["acf"]["blocks"][0]["free_form_content"]["content"]
            process_content_to_insert(content, country_name)
        else:
            for i in range(0, govt_count):
                content = json_data["result"]["data"]["page"]["acf"]["blocks"][i]["free_form_content"]["content"]
                process_content_to_insert(content, country_name)


def fetch_designations(country_name):
    des_list = list()
    check_designation = "SELECT designation from world_leaders where country = %s"
    designations = dc.execute_select_query(check_designation, (country_name,))
    for designation in designations:
        des_list.append(designation[0])
    return des_list


def convert_to_pep_import_df(data):
    return pd.DataFrame(data, columns=['country', 'designation', 'leader_name'])


def fetch_world_leaders_history():
    data = dc.execute_select_query("SELECT country, designation, leader_name from wl_import_history", ())
    return convert_to_pep_import_df(data)


def process_world_leaders():
    result_set = dc.execute_select_query("SELECT id, country, designation, leader_name FROM wl_import_history WHERE deleted_on is null", ())
    for result in result_set:
        import_id = result[0]
        country = result[1]
        designation = result[2]
        leader_name = clean_ignore_words(result[3])
        do_entry_in_db(country, designation, leader_name, import_id)


def do_import_history_entry(country_name, designation, leader_name):
    insert_qry = "INSERT INTO wl_import_history(country, designation, leader_name, created_on, deleted_on) VALUES (%s, %s, %s, %s, %s)"
    duplication_qry = "SELECT * from wl_import_history where country = %s and designation = %s and leader_name = %s"
    insert_val = (country_name, designation, leader_name, curr_timestamp, None)
    if not dc.check_for_duplication(duplication_qry, (country_name, designation, leader_name)):
        dc.execute_query_single(insert_qry, insert_val)


def do_entry_in_db(country_name, designation, leader_name, import_id):
    insert_qry = "INSERT INTO world_leaders(country, designation, leader_name, import_id, created_on, deleted_on) VALUES (%s, %s, %s, %s, %s, %s)"
    check_duplication_qry = "SELECT * from world_leaders where country = %s and designation = %s and leader_name = %s"
    insert_val = (country_name, designation, leader_name, import_id, curr_timestamp, None)
    if not dc.check_for_duplication(check_duplication_qry, (country_name, designation, leader_name)):
        dc.execute_query_single(insert_qry, insert_val)
        # if designation in fetch_designations(country_name):  # check if this designation exists for this country
        #     leader = dc.execute_select_query(check_leader, leader_qry_val)
        #     if leader is not [] and leader[0][0] != leader_name:  # check if leader is changed for this designation
        #         dc.execute_query_single(update_leader, (curr_timestamp, leader_name, country_name, designation))
        # else:  # no such entry, so do a new entry
        #     dc.execute_query_single(insert_qry, insert_val)


def process_content(content, country_name):
    bs_data = bS(content, "lxml-html")
    data = []
    for designation_tag in bs_data.find_all("h3"):
        designation = clean_text(designation_tag.text.strip())
        if designation_tag.find_next() is not None and designation_tag.find_next().name == "p":
            leader_name = clean_text(designation_tag.find_next().text.strip())
        else:
            leader_name = ""
        data.append([country_name, designation, leader_name])
    return data


def process_content_to_insert(content, country_name):
    bs_data = bS(content, "lxml-html")
    # data = []
    for designation_tag in bs_data.find_all("h3"):
        designation = clean_text(designation_tag.text.strip())
        if designation_tag.find_next() is not None and designation_tag.find_next().name == "p":
            leader_name = clean_text(designation_tag.find_next().text.strip())
        else:
            leader_name = ""
        # data.append([country_name, designation, leader_name])
    # return data
        do_import_history_entry(country_name, designation, leader_name)


def update_wl_import(prev_data, new_data):
    df = pd.merge(new_data, prev_data, on=['country', 'designation', 'leader_name'], how="left", indicator=True)
    records_to_be_added = df[df['_merge'] == 'left_only']
    records_to_be_added = records_to_be_added.drop('_merge', 1)
    # add_count = records_to_be_added.count()

    df = pd.merge(prev_data, new_data, on=['country', 'designation', 'leader_name'], how="left", indicator=True)
    records_to_be_deleted = df[df['_merge'] == 'left_only']
    records_to_be_deleted = records_to_be_deleted.drop('_merge', 1)
    # deleted_count = records_to_be_deleted.count()

    now = get_current_timestamp()
    add_query = 'INSERT INTO wl_import_history(country, designation, leader_name, created_on) VALUES (%s, %s, %s, %s)'
    for index, row in records_to_be_added.iterrows():
        dc.execute_query_single(add_query, (row['country'], row['designation'], row['leader_name'], now))

    delete_query = "UPDATE wl_import_history set deleted_on = '" + now + "' where country = %s and designation = %s and leader_name = %s"
    for index, row in records_to_be_deleted.iterrows():
        dc.execute_query_single(delete_query, (row['country'], row['designation'], row['leader_name']))
    # return add_count, deleted_count


def updated_deleted_wl():
    deleted_import_ids = dc.execute_select_query("select id from wl_import_history where deleted_on is not null", ())
    print(deleted_import_ids)
    if deleted_import_ids is not [] and len(deleted_import_ids) > 0:
        ids = []
        for row in deleted_import_ids:
            ids.append("import_id = " + str(row[0]))
        now = get_current_timestamp()
        dc.execute_query_single('update world_leaders set deleted_on = %s where ' + ' OR '.join(ids), (now,))


def process_excel():
    result_set = dc.execute_select_query("SELECT UPPER(country) AS 'COUNTRY', designation AS 'DESIGNATION', leader_name AS 'LEADER NAME', IF(deleted_on is null, 'ACTIVE', 'INACTIVE') AS STATUS, created_on AS 'CREATED ON', deleted_on AS 'DELETED ON' FROM world_leaders", ())
    df = pd.DataFrame(result_set, columns=['COUNTRY', 'DESIGNATION', 'LEADER NAME', 'STATUS', 'CREATED ON', 'DELETED ON'])
    if df is not None:
        generate_excel_from_dataframe(df, ws)
        wb.save(filename=dest_filename)


def check_entries_in_wl_history():
    return len(dc.execute_select_query("SELECT * from wl_import_history", ()))


if check_entries_in_wl_history() == 0:
    parse_json_file_to_insert()
    process_world_leaders()
else:
    prev_data = fetch_world_leaders_history()
    new_data = parse_json_file()
    update_wl_import(prev_data, new_data)
    process_world_leaders()
    updated_deleted_wl()

process_excel()
