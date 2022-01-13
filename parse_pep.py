import datetime
import json
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup as bS
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font
import database_operations as dc
import environment
import time
import email_utility as eu


db = dc.connect_db()
curr_timestamp = datetime.datetime.now().replace(microsecond=0)
today = str(datetime.date.today())  # - datetime.timedelta(days=1))
url = "https://www.cia.gov/the-world-factbook/page-data/field/political-parties-and-leaders/page-data.json"

wb = Workbook()
wb1 = Workbook()
email_for_manual_entries = environment.email_ids_for_pep_sme
manual_entries_filename = environment.export_dir + 'pep/manual_entries.xlsx'
pep_dest_filename = 'pep/political_parties_and_leaders.xlsx'
ws1 = wb.active
ws1.title = "Full List of all PEP"
ws2 = wb.create_sheet("Diff (Previous Day Changes)")
ws3 = wb1.active
ws3.title = 'Manual Entries'
manual = "MANUAL"
automated = "AUTOMATED"

def fetch_row(import_id):
    return dc.get_entity("SELECT row_content from pep_import_history where id = %s", (import_id,))

def generate_excel_from_dataframe(df, work_sheet, highlight=False):
    work_sheet.append(tuple(df.keys()))
    header_len = len(tuple(df.keys()))
    for i in range(1, header_len + 1):
        cell = chr(i + 64) + str(1)
        work_sheet[cell].font = Font(bold=True)
    if not highlight:
        for tup in df.itertuples(index=False):
            work_sheet.append(tup)
    else:
        my_red = Color(rgb='00FFCCCB')
        my_fill = PatternFill(patternType='solid', fgColor=my_red)
        for index, tup in enumerate(df.itertuples(index=False)):
            row = tup[1]
            work_sheet.append(tup)
            if not ("[" in row and "]" in row):
                for i in range(1, header_len+1):
                    cell = chr(i + 64) + str(index+2)
                    work_sheet[cell].fill = my_fill

def process_sheet1():
    result_set = dc.execute_select_query("SELECT country, party_name, leader_name FROM pep where deleted_on = null", ())
    country_list = list()
    party_list = list()
    leader_list = list()
    for result in result_set:
        country_list.append(result[0])
        party_list.append(result[1])
        leader_list.append(result[2])
    data = {'Country': country_list, 'Party_Name': party_list, 'Leader_Name': leader_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws1)


def process_sheet2():
    result_set = dc.execute_select_query("SELECT country, party_name, leader_name, import_id, deleted_on FROM pep", ())
    country_list = list()
    party_list = list()
    leader_list = list()
    row_list = list()
    action_list = list()
    for result in result_set:
        country_list.append(result[0])
        party_list.append(result[1])
        leader_list.append(result[2])
        import_id = result[3]
        text_as_is = fetch_row(import_id)
        row_list.append(text_as_is[0])
        if result[4] is not None:
            action_list.append("DELETED")
        else:
            action_list.append("ADDED")
    data = {'Text_As_Is': row_list, 'Country': country_list, 'Party_Name': party_list,
            'Leader_Name': leader_list, 'Action': action_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws2)
    wb.save(filename=pep_dest_filename)
    # eu.initiate_email(emails_for_pep_data, subject, content, pep_dest_filename)


def notify_manual_entries_to_sme():
    result_set = dc.execute_select_query("SELECT country, row_content, is_note FROM pep_import_history WHERE type = 'MANUAL' AND created_on > '"+today+" 00:00:00'", ())
    country_list = list()
    row_list = list()
    is_note_list = list()
    if result_set == None or len(result_set) == 0:
        return 0
    for result in result_set:
        country_list.append(result[0])
        row_list.append(result[1])
        is_note_list.append('Yes' if result[2] == 1 else 'No')
    data = {'Country': country_list, 'Row': row_list, 'Is_It_A_Note': is_note_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws3, True)
    subject = "Manual Entries in PEP List"
    content = "Hello SME Team, \n\nAttached excel contains the manual entries to be filtered, " \
              "validated & inserted into the pep table in database."
    wb1.save(filename=manual_entries_filename)
    eu.initiate_email(email_for_manual_entries, subject, content, manual_entries_filename)
    return len(result_set)


def clean_text_with_bs(text):
    if not text.endswith(">"):
        remaining_text = text.split(">")[len(text.split(">")) - 1]
        html_data = bS(text, "lxml-html")
        if html_data.text == remaining_text or remaining_text in html_data.text:
            return_text = html_data.text.strip()
        else:
            return_text = (html_data.text + remaining_text).strip()
    else:
        html_data = bS(text, "lxml-xml")
        return_text = html_data.text.strip()
    if return_text == "":
        return text.replace("<strong>", "").strip()
    else:
        return return_text


# call the URL & collect response
def fetch_data_from_url():
    file = urllib.request.urlopen(url)
    data = file.read()
    file.close()
    return str(data, "UTF-8")


# store the response into an XML file in string format
def do_file_operations():
    file_name = environment.dir_name + "pep/" + today + ".json"
    f = open(file_name, "w")
    f.write(fetch_data_from_url())
    f.close()
    return file_name


def check_for_keywords(data_string):
    select_keywords_qry = "SELECT * from pep_import_keywords where keyword = %s and is_ignored = %s"
    return dc.check_for_duplication(select_keywords_qry, (data_string, 1))


def get_designations():
    designations = dc.execute_select_query("SELECT designation from pep_designations where 1", ())
    return designations


def check_for_designation_string(leader):
    designations = get_designations()
    des_list = list()
    for designation in designations:
        des_list.append(designation[0])
    for designation in des_list:
        leader = leader.replace(designation, "").strip()
    return leader


def store_pep_history(country_name, row, entry_type, is_note, created_on, updated_on, deleted_on):
    pep_hist_insert = "INSERT INTO pep_import_history" \
                      "(country, row_content, type, is_note, created_on, modified_on, deleted_on) " \
                      "VALUES (%s, %s, %s, %s, %s, %s, %s)"
    duplication_qry = "SELECT * from pep_import_history where country = %s and row_content = %s and type = %s and is_note = %s"
    insert_values = (country_name, row, entry_type, is_note, created_on, updated_on, deleted_on)
    check_values = (country_name, row, entry_type, is_note)
    if not dc.check_for_duplication(duplication_qry, check_values):
        dc.execute_query_single(pep_hist_insert, insert_values)


# def check_comma_after_bracket(data_string):
#     if "]" in data_string:
#         splits = data_string.split("]")
#         for split in splits:
#             if split.startswith(","):
#                 return True
#     else:
#         return False


def process_pep_import(country_name, data_string):
    if "<" in data_string:
        data_string = clean_text_with_bs(data_string.replace("<p>", "").replace("</p>", "")).replace("</strong>", "").strip()
    left_bracket_count = data_string.count('[')
    right_bracket_count = data_string.count(']')
    if left_bracket_count == 1 and right_bracket_count == 1:
        if data_string.split("]")[1] != "" and data_string.split("]")[1].startswith(")"):
            return [country_name, data_string, manual, False]
        elif data_string.startswith("note:") or data_string.startswith("note :"):
            is_note = True if check_for_keywords(data_string) else False
            return [country_name, data_string, manual, is_note]
        elif data_string.startswith("one registered party: ") and data_string.replace("one registered party: ", "") != "":
            return [country_name, data_string.replace("one registered party: ", ""), automated, False]
        elif data_string.startswith("other: ") and data_string.replace("other: ", "") != "":
            return [country_name, data_string.replace("other: ", ""), automated, False]
        elif data_string.startswith("New - ") and data_string.replace("New - ", "") != "":
            return [country_name, data_string.replace("New - ", ""), automated, False]
        elif data_string.split("[")[1].replace("]", "").startswith("collective leadership") \
                or data_string.split("[")[1].replace("]", "") == "NA" \
                or data_string.split("[")[1].replace("]", "") == "N/A"\
                or data_string.split("[")[1].replace("]", "") == "joint leadership of several medical doctors":
            pass
        else:
            return [country_name, data_string, automated, False]
    else:
        if data_string != "" and data_string != "(" and (not (data_string.startswith("<") and data_string.endswith(">"))):
            is_note = True if check_for_keywords(data_string) else False
            return [country_name, data_string, manual, is_note]
    return [country_name, data_string, manual, False]
        
def process_pep_history(country_name, data_string, mark_as_deleted):
    today_date = datetime.date.today()
    created_on = today_date - datetime.timedelta(days=1) if mark_as_deleted else today_date
    deleted_on = today_date if mark_as_deleted else None
    if "<" in data_string:
        data_string = clean_text_with_bs(data_string.replace("<p>", "").replace("</p>", "")).replace("</strong>", "").strip()
    left_bracket_count = data_string.count('[')
    right_bracket_count = data_string.count(']')
    if left_bracket_count == 1 and right_bracket_count == 1:
        if data_string.split("]")[1] != "" and data_string.split("]")[1].startswith(")"):
            store_pep_history(country_name, data_string, manual, False, created_on, None, deleted_on)
        elif data_string.startswith("note:") or data_string.startswith("note :"):
            is_note = True if check_for_keywords(data_string) else False
            store_pep_history(country_name, data_string, manual, is_note, created_on, None, deleted_on)
        elif data_string.startswith("one registered party: ") and data_string.replace("one registered party: ", "") != "":
            store_pep_history(country_name, data_string.replace("one registered party: ", ""), automated, False, created_on, None, deleted_on)
        elif data_string.startswith("other: ") and data_string.replace("other: ", "") != "":
            store_pep_history(country_name, data_string.replace("other: ", ""), automated, False, created_on, None, deleted_on)
        elif data_string.startswith("New - ") and data_string.replace("New - ", "") != "":
            store_pep_history(country_name, data_string.replace("New - ", ""), automated, False, created_on, None, deleted_on)
        elif data_string.split("[")[1].replace("]", "").startswith("collective leadership") \
                or data_string.split("[")[1].replace("]", "") == "NA" \
                or data_string.split("[")[1].replace("]", "") == "N/A"\
                or data_string.split("[")[1].replace("]", "") == "joint leadership of several medical doctors":
            pass
        else:
            store_pep_history(country_name, data_string, automated, False, created_on, None, deleted_on)
    else:
        if data_string != "" and data_string != "(" and (not (data_string.startswith("<") and data_string.endswith(">"))):
            is_note = True if check_for_keywords(data_string) else False
            store_pep_history(country_name, data_string, manual, is_note, created_on, None, deleted_on)
        # elif check_comma_after_bracket(data_string): # case for Uruguay 1st merged entry
        #     for split in data_string.split("],"):
        #         store_pep_history(country_name, split, automated, False, today_date, today_date, None)

def convert_to_pep_import_df(data):
    return pd.DataFrame(data, columns = ['country', 'row_content', 'type', 'is_note'])

def fetch_pep_history():
    data = dc.execute_select_query("SELECT country, row_content, type, is_note from pep_import_history", ())
    return convert_to_pep_import_df(data)

def process_data(data_string):
    note = ""
    if "[" in data_string:
        party_name = data_string.split("[")[0]
        party_leader = data_string.split("[")[1]
        if "]" not in party_leader:
            party_leader += "]"
        if not party_leader.split("]")[1] == "":
            note = party_leader.split("]")[1].strip()
            party_leader = party_leader.replace(note, "").strip()
            if note.startswith("<br><br>"):
                note = ""
    elif "(" in data_string:
        party_name = data_string.split("(")[0]
        party_leader = ""
        note = "(" + data_string.split("(")[1]
        # splitted_data = note.replace("(", "").replace(")", "").split(" ")
        if note.replace("(", "").replace(")", "").isupper():  # for Azerbaijan -> (PDR) scenario
            party_name += note
            note = ""
        # elif len(splitted_data) > 1 and splitted_data[
        #     1].isupper():  # for scenarios where party leaders names are in () parenthesis & not in []
        #     party_leader = note.replace("(", "").replace(")", "")
        #     note = ""
    elif "; note" in data_string or ";note" in data_string:
        if "; note" in data_string:
            party_name = data_string.split("; note")[0]
            party_leader = ""
            note = data_string.split("; note")[1]
        else:
            party_name = data_string.split(";note")[0]
            party_leader = ""
            note = data_string.split(";note")[1]
    else:
        party_name = data_string
        party_leader = ""
        note = ""
    data_tuple = (party_name.strip(),
                  party_leader.replace("]", "").replace(note.strip(), "").strip(),
                  note.lstrip(";").strip())
    return data_tuple


def clean_note_string(note):
    if note == "," or note == ")" or note == "(":
        note = ""
    return note.replace("&nbsp;", " ").replace("<p>", "").replace("</p>", "").replace("<br>", "") \
        .replace("; note", "").replace(";note", "").replace("&amp;", "&").replace("&rsquo;", "’") \
        .replace("<strong>", "").replace("</strong>", "").strip().lstrip("-").lstrip(";")


def clean_party_name_string(party_name):
    return party_name.replace("&amp;", "&").replace("&nbsp;", "").replace("&rsquo;", "’").strip()


def clean_party_leader_string(party_leader):
    if party_leader.endswith(")"):
        party_leader = party_leader.replace(")", "").strip()
    return party_leader.replace("&amp;", "&").replace("&nbsp;", " ").replace("&rsquo;", "’")\
        .replace("&aacute;", "á").replace(" (both claiming leadership)", "").strip()


def parse_json_file():
    file_name = environment.dir_name + "pep/" + today + ".json"
    f = open(file_name,)
    json_data = json.load(f)
    political_parties_json_string = json_data["result"]["data"]["page"]["json"].replace("\"", '"')
    pp_json_data = json.loads(political_parties_json_string)
    data = []
    for country_data in pp_json_data["countries"]:
        country_name = country_data["name"].strip()
        parties_string = country_data["data"].replace("&amp;", "&").replace("&nbsp;", " ").replace("&rsquo;", "’").strip()
        if "<br />" in parties_string:
            parties = parties_string.split("<br />")
            for party in parties:
                if party != "":
                    party = party.replace("[[", "[").replace("]]", "]").strip()
                    if "<br>" in party:
                        if "<br><br>" in party:
                            first_part = party.split("<br><br>")[0]
                            second_part = party.split("<br><br>")[1]
                        else:
                            first_part = party.split("<br>")[0]
                            second_part = party.split("<br>")[1]
                        if first_part.strip() != "":
                            data.append(process_pep_import(country_name, first_part))
                        if second_part.strip() != "":
                            data.append(process_pep_import(country_name, second_part))
                    else:
                        data.append(process_pep_import(country_name, party))
        elif "<br>" in parties_string:
            parties = parties_string.split("<br />")
            party = parties[0].replace("[[", "[").replace("]]", "]").strip()
            if party != "":
                if "<br>" in party:
                    if "<br><br>" in party:
                        first_part = party.split("<br><br>")[0]
                        second_part = party.split("<br><br>")[1]
                    else:
                        first_part = party.split("<br>")[0]
                        second_part = party.split("<br>")[1]
                    if first_part.strip() != "":
                        data.append(process_pep_import(country_name, first_part))
                    if second_part.strip() != "":
                        data.append(process_pep_import(country_name, second_part))
                else:
                    data.append(process_pep_import(country_name, party))
        else:
            parties_string = parties_string.replace("[[", "[").replace("]]", "]").strip()
            if parties_string != "":
                data.append(process_pep_import(country_name, parties_string))
    return convert_to_pep_import_df(data)

def process_automated_peps():
    result_set = dc.execute_select_query("SELECT id, country, row_content FROM pep_import_history WHERE type = 'AUTOMATED' and deleted_on is null", ())
    for result in result_set:
        import_id = result[0]
        country = result[1]
        row_content = result[2]
        data_tuple = process_data(row_content)
        party = data_tuple[0]
        party_leaders = data_tuple[1].replace("coalition led by ", "").replace("associated with former ", "")\
            .replace("Central Committee", "").strip()
        note = data_tuple[2]
        if "+" in party_leaders:
            party_leader_splitted = party_leaders.split("+")
            for leader in party_leader_splitted:
                leader_name = check_for_designation_string(leader.strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, None, note, automated, import_id)
        elif "," in party_leaders:
            party_leader_splitted = party_leaders.split(",")
            for index, leader in enumerate(party_leader_splitted):
                if " and " in leader:
                    party_leader_splitted1 = leader.split(" and ")
                    for leader1 in party_leader_splitted1:
                        leader_name = check_for_designation_string(leader1.strip()).lstrip("-")
                        if leader_name != "" and leader_name.strip() != "Dr." and leader_name != "Jr.":
                            do_entry_in_pep(country, party, leader_name, None, note, automated, import_id)
                else:
                    if not leader.strip().startswith("aka"):
                        leader_alias = None
                        if index <= 0:
                            if party_leader_splitted[index+1] is not None:
                                if party_leader_splitted[index+1].strip().startswith("aka"):
                                    leader_alias = party_leader_splitted[index+1].strip().replace("aka", "")
                            leader_name = check_for_designation_string(leader.strip()).lstrip("-")
                            if leader_name != "" and leader_name.strip() != "Dr." and leader_name.strip() != "Jr.":
                                do_entry_in_pep(country, party, leader_name, leader_alias, note, automated, import_id)
                        else:
                            leader_name = check_for_designation_string(leader.strip()).lstrip("-")
                            if leader_name != "" and leader_name.strip() != "Dr." and leader_name.strip() != "Jr.":
                                do_entry_in_pep(country, party, leader_name, leader_alias, note, automated, import_id)
        elif " and " in party_leaders:
            party_leader_splitted = party_leaders.split(" and ")
            for leader in party_leader_splitted:
                leader_name = check_for_designation_string(leader.replace("MPs", "").strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, None, note, automated, import_id)
        elif ";" in party_leaders:
            party_leader_splitted = party_leaders.split(";")
            for leader in party_leader_splitted:
                leader_name = check_for_designation_string(leader.strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, None, note, automated, import_id)
        elif " or " in party_leaders:
            party_leader_splitted = party_leaders.split(" or ")
            if len(party_leader_splitted) == 2:
                leader = party_leader_splitted[0]
                leader_alias = party_leader_splitted[1]
                leader_name = check_for_designation_string(leader.strip())
                if leader_name != "":
                    do_entry_in_pep(country, party, leader_name, leader_alias, note, automated, import_id)
        else:
            leader_name = check_for_designation_string(party_leaders.strip())
            if leader_name != "":
                do_entry_in_pep(country, party, leader_name, None, note, automated, import_id)


def do_entry_in_pep(country_name, party_name, leaders, leader_alias, notes, entry_type, import_id):
    check_duplication_qry = "SELECT * from pep where country = %s and party_name = %s and leader_name = %s " \
                            "and notes = %s and type = %s and import_id = %s and deleted_on is null"
    insertion_qry = "INSERT INTO pep(country, party_name, leader_name, leader_alias, notes, type, import_id, " \
                    "created_on) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
    party_name = clean_party_name_string(party_name)
    leaders = clean_party_leader_string(leaders)
    notes = clean_note_string(notes)
    now = get_current_timestamp()
    if not (party_name == "" and leaders == "" and notes == ""):
        dup_val = (country_name, party_name, leaders, notes, entry_type, import_id)
        if not dc.check_for_duplication(check_duplication_qry, dup_val):
            insert_val = (country_name, party_name, leaders, leader_alias, notes, entry_type, import_id, now)
            dc.execute_query_single(insertion_qry, insert_val)


def updated_deleted_pep():
    deleted_import_ids = dc.execute_query_single("select id from pep_import_history where deleted_on is not null", ())
    if deleted_import_ids != None and len(deleted_import_ids) > 0:
        ids = []
        for row in deleted_import_ids:
            ids.append("import_id = " + str(row[0]))
        now = get_current_timestamp()
        dc.execute_query_single('update pep set deleted_on = %s where ' + ' OR '.join(ids), (now,))


def check_row_deletion(country_data, new_data):
    for data in new_data:  # (A, B, C)
        if data[0] == country_data[0]:  # D
            return False
    return True


def check_row_addition(country_data, prev_data):
    for data in prev_data:
        if data[0] == country_data[0]:
            return True
    return False


def process_pep_changes(prev_data, new_data):
    prev_countries = prev_data.keys()
    new_countries = new_data.keys()
    if len(prev_countries) == len(new_countries):
        for country in prev_countries:
            print(country)
            if len(prev_data[country]) > len(new_data[country]):  # some rows got deleted from recent data
                for country_data in prev_data[country]:
                    if check_row_deletion(country_data, new_data[country]):
                        process_pep_history(country, country_data[0], True)
            # elif len(prev_data[country]) < len(new_data[country]):  # some rows got added to recent data
            #     for country_data in new_data[country]:
            #         if check_row_addition(country_data, prev_data[country]):
            #             process_pep_history(country, country_data[0], False)

def get_current_timestamp():
    ts = time.time()
    return datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') 

def update_pep_import(prev_data, new_data):
    df = pd.merge(new_data, prev_data, on=['country','row_content', 'type', 'is_note'], how="left", indicator=True)
    records_to_be_added = df[df['_merge'] == 'left_only']
    records_to_be_added = records_to_be_added.drop('_merge', 1)
    new_manual_records_count = len(records_to_be_added[records_to_be_added['type'] == 'MANUAL'])
    new_automated_records_count = len(records_to_be_added[records_to_be_added['type'] == 'AUTOMATED'])
    
    df = pd.merge(prev_data, new_data, on=['country','row_content', 'type', 'is_note'], how="left", indicator=True)
    records_to_be_deleted = df[df['_merge'] == 'left_only']
    records_to_be_deleted = records_to_be_deleted.drop('_merge', 1)
    deleted_manual_records_count = len(records_to_be_deleted[records_to_be_deleted['type'] == 'MANUAL'])
    deleted_automated_records_count = len(records_to_be_deleted[records_to_be_deleted['type'] == 'AUTOMATED'])
    
    now = get_current_timestamp()
    add_query = 'INSERT INTO pep_import_history (country, row_content, type, is_note, created_on) VALUES (%s, %s, %s, %s, %s)'
    for index, row in records_to_be_added.iterrows():
        dc.execute_query_single(add_query, (row['country'], row['row_content'], row['type'], row['is_note'], now))
    
    delete_query = "UPDATE pep_import_history set deleted_on = '" + now +"' where country = %s and row_content = %s and type = %s and is_note = %s "
    for index, row in records_to_be_deleted.iterrows():
        dc.execute_query_single(delete_query, (row['country'], row['row_content'], row['type'], row['is_note']))
        
    return new_manual_records_count, new_automated_records_count, deleted_manual_records_count, deleted_automated_records_count

def audit_parse_report(new_manual_records_count, new_automated_records_count, deleted_manual_records_count, deleted_automated_records_count):
    is_manual_rectified = 0
    if new_manual_records_count > 0:
        action = "PENDING_RECTIFICATION"
    elif new_manual_records_count == 0 and new_automated_records_count == 0 and deleted_manual_records_count == 0 and deleted_automated_records_count == 0:
        action = "NO_CHANGE_REPORTED"
    else:
        action = "REPORT_CHANGE"
        
    query = "INSERT INTO pep_notification_log (action, date, is_manual_rectified, count_new_automated, count_new_manual, count_deleted_automated, count_deleted_manual) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    dc.execute_query_single(query, (action, today, is_manual_rectified, new_automated_records_count, new_manual_records_count, deleted_automated_records_count, deleted_manual_records_count))
    
do_file_operations()

prev_data = fetch_pep_history()  
new_data = parse_json_file() 

new_manual_records_count, new_automated_records_count, deleted_manual_records_count, deleted_automated_records_count = update_pep_import(prev_data, new_data)

updated_deleted_pep()

process_automated_peps()

notify_manual_entries_to_sme() 

audit_parse_report(new_manual_records_count, new_automated_records_count, deleted_manual_records_count, deleted_automated_records_count)