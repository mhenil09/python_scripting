import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font
import database_operations as dc
import email_utility as eu
import environment

db = dc.connect_db()
curr_timestamp = datetime.datetime.now().replace(microsecond=0)
today = str(datetime.date.today()) 

wb = Workbook()
emails_for_pep_data = environment.emails_ids_for_pep_export
pep_dest_filename = environment.export_dir + 'pep/political_parties_and_leaders.xlsx'
ws1 = wb.active
ws1.title = "Full List of all PEP"
ws2 = wb.create_sheet("Diff (Previous Day Changes)")

def fetch_row(import_id):
    return dc.get_entity("SELECT row_content from pep_import_history where id = %s", (import_id,))

def generate_excel_from_dataframe(df, work_sheet, highlight=False):
    work_sheet.append(tuple(df.keys()))
    header_len = len(tuple(df.keys()))
    for i in range(1, header_len + 1):
        cell = chr(i + 64) + str(1)
        work_sheet[cell].font = Font(bold=True)
    if not highlight:
        for tup in df.itertuples(index=False):
            work_sheet.append(tup)
    else:
        my_red = Color(rgb='00FFCCCB')
        my_fill = PatternFill(patternType='solid', fgColor=my_red)
        for index, tup in enumerate(df.itertuples(index=False)):
            row = tup[1]
            work_sheet.append(tup)
            if not ("[" in row and "]" in row):
                for i in range(1, header_len+1):
                    cell = chr(i + 64) + str(index+2)
                    work_sheet[cell].fill = my_fill

def process_sheet1():
    result_set = dc.execute_select_query("SELECT country, party_name, leader_name, created_on, deleted_on FROM pep", ())
    country_list = list()
    party_list = list()
    leader_list = list()
    created_dates = list()
    deleted_dates = list()
    for result in result_set:
        country_list.append(result[0])
        party_list.append(result[1])
        leader_list.append(result[2])
        created_dates.append(result[3])
        deleted_dates.append(result[4])
    data = {'Country': country_list, 'Party_Name': party_list, 'Leader_Name': leader_list, 'Created_On': created_dates, 'Deleted_On': deleted_dates}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws1)

def process_sheet2():
    result_set = dc.execute_select_query("SELECT country, party_name, leader_name, import_id, deleted_on FROM pep where deleted_on > '"+today+" 00:00:00' or created_on > '"+today+" 00:00:00'", ())
    country_list = list()
    party_list = list()
    leader_list = list()
    row_list = list()
    action_list = list()
    for result in result_set:
        country_list.append(result[0])
        party_list.append(result[1])
        leader_list.append(result[2])
        import_id = result[3]
        text_as_is = fetch_row(import_id)
        row_list.append(text_as_is[0])
        if result[4] is not None:
            action_list.append("DELETED")
        else:
            action_list.append("ADDED")
    data = {'Text_As_Is': row_list, 'Country': country_list, 'Party_Name': party_list,
            'Leader_Name': leader_list, 'Action': action_list}
    df = pd.DataFrame(data)
    if df is not None:
        generate_excel_from_dataframe(df, ws2)
    subject = 'PEP data'
    content = 'Hello Team,\n\nAttached excel contains PEP data in 2 versions' \
              '\n\n\n   1. Full List of all PEP\n   2. Diff (Previous Day Changes)'
    wb.save(filename=pep_dest_filename)
    eu.initiate_email(emails_for_pep_data, subject, content, pep_dest_filename)

def send_report():
    query = "SELECT id, action FROM pep_notification_log where date = %s"
    result = dc.execute_select_query(query, (today,))
    if result != None and len(result) > 0:
        action = result[0][1]
        if action == 'REPORT_CHANGE':
            return True, result[0][0]
    return False, None

def update_pep_notification_log(log_id):
    query = "update pep_notification_log set action = 'CHANGE_REPORTED' where id = " + str(log_id) 
    print(query)
    dc.execute_select_query(query, ())

send_report_flag, log_id = send_report()
if send_report_flag:
    process_sheet1()   
    process_sheet2()
    update_pep_notification_log(log_id)